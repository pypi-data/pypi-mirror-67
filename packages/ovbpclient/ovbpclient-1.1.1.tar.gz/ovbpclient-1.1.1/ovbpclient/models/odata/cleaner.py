import os
import logging

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.formatting.rule import FormulaRule

import ovbpclient
from ..base import BaseModel
from .mixin_generator import GeneratorModelMixin
from ...exceptions import RecordDoesNotExistError
from ...util import get_one_and_only_one

module_path = os.path.realpath(os.path.join(ovbpclient.__file__, ".."))

logger = logging.getLogger(__name__)


class Cleaner(BaseModel, GeneratorModelMixin):
    def get_odata_project(self):
        return self.client.odata_projects.retrieve(self.project)

    def get_unitcleaner(self, external_name):
        unitcleaners_l = self.client.unitcleaners.list(
            filter_by=dict(cleaner=self.id, external_name=external_name),
            limit=2)
        return get_one_and_only_one(unitcleaners_l)

    def create_unitcleaner(
            self,
            external_name,
            name,
            freq="1H",
            input_unit_type="instantaneous",
            unit_type="instantaneous",
            input_convention="left",
            clock="tzt",
            timezone="Europe/Paris",
            unit="",
            resample_rule="mean",
            interpolate_limit=0,
            wait_offset="6H",
            label="",
            input_expected_regular=False,
            operation_fct=None,
            filter_fct=None,
            derivative_filter_fct=None,
            custom_delay=None,
            custom_fct=None,
            custom_before_offset=None,
            custom_after_offset=None
    ):
        data = dict(
            cleaner=self.id,
            external_name=external_name,
            name=name,
            freq=freq,
            input_unit_type=input_unit_type,
            unit_type=unit_type,
            input_convention=input_convention,
            clock=clock,
            timezone=timezone,
            unit=unit,
            resample_rule=resample_rule,
            interpolate_limit=interpolate_limit,
            wait_offset=wait_offset,
            label=label,
            input_expected_regular=input_expected_regular,
            operation_fct=operation_fct,
            filter_fct=filter_fct,
            derivative_filter_fct=derivative_filter_fct,
            custom_delay=custom_delay,
            custom_fct=custom_fct,
            custom_before_offset=custom_before_offset,
            custom_after_offset=custom_after_offset
        )
        return self.client.unitcleaners.create(**data)

    def list_all_importer_series(self):
        return self.client.importer_series.list_all(filter_by=dict(generator=self.related_importer))

    def list_all_unitcleaners(self):
        return self.client.unitcleaners.list_all(filter_by=dict(cleaner=self.id))

    def configure_from_excel(self, path):
        # retrieve current unitcleaners data
        unitcleaners = self.list_all_unitcleaners()

        # load excel data
        all_excel_unitcleaners_data = excel_to_unitcleaners_data(path)

        # iter excel data
        for excel_unitcleaner_data in all_excel_unitcleaners_data:
            # see if unitcleaner exists
            unitcleaner = None
            try:
                unitcleaner = get_one_and_only_one(
                    unitcleaners,
                    lambda x: x.external_name == excel_unitcleaner_data["external_name"]
                )
            except RecordDoesNotExistError:
                pass

            # manage both workflows
            # 1. already exists
            if unitcleaner is not None:
                diff = dict()
                for k, v in excel_unitcleaner_data.items():
                    if getattr(unitcleaner, k) != v:
                        diff[k] = v
                # if diff, update
                if len(diff) > 0:
                    logging.info(f"unitcleaner {unitcleaner.name} already exists, updating")
                    unitcleaner.update(**diff)
                else:
                    logging.info(f"unitcleaner {unitcleaner.name} has not changed, skipping update")

            # 2. does not exist
            else:
                # create
                self.create_unitcleaner(**excel_unitcleaner_data)

    def export_configuration_to_excel(self, path):
        odata_project = self.get_odata_project()
        cleaner_data = self.data
        unitcleaners = self.list_all_unitcleaners()
        not_configured_series = [se for se in self.list_all_importer_series() if se.unitcleaner is None]

        configured_nb = len(unitcleaners)
        non_configured_nb = len(not_configured_series)
        tot_nb = configured_nb + non_configured_nb
        start_row = 8
        last_row = tot_nb + start_row - 1
        start_col = 1
        last_col = 23

        # *******************
        # Load model workbook
        # *******************

        wb = load_workbook(os.path.join(module_path, "resources", 'multiclean_config_model.xlsx'))
        ws = wb["Series"]

        ws['B2'].value = odata_project.name
        ws['B3'].value = cleaner_data['name']

        # ****************
        # Validation Lists
        # ****************

        validation = dict(
            save_config=dict(
                dv=DataValidation(
                    type="list",
                    formula1='"x"',
                    allow_blank=True),
                col="C"
            ),
            active=dict(
                dv=DataValidation(
                    type="list",
                    formula1='"yes, no"',
                    allow_blank=True),
                col="D"
            ),
            input_ext=dict(
                dv=DataValidation(
                    type="list",
                    formula1='"instantaneous, delta, mean, bonjour"',
                    allow_blank=True),
                col="F"
            ),
            convention=dict(
                dv=DataValidation(type="list", formula1='"left, right"', allow_blank=True),
                col="G"
            ),
            clock=dict(
                dv=DataValidation(type="list", formula1='"dst, gmt, tzt"', allow_blank=True),
                col="H"
            ),
            timezone=dict(
                dv=DataValidation(
                    type="list",
                    formula1='"Europe/Paris, Africa/Abidjan, Africa/Accra"',
                    allow_blank=True),
                col="I"
            ),
            input_is_regular=dict(
                dv=DataValidation(
                    type="list",
                    formula1='"yes, no"',
                    allow_blank=True),
                col="L"
            ),
            output_ext=dict(
                dv=DataValidation(
                    type="list",
                    formula1='"instantaneous, delta, mean, mean_derivate"',
                    allow_blank=True),
                col="M"
            ),
            reasmple=dict(
                dv=DataValidation(
                    type="list",
                    formula1='"mean, sum, first"',
                    allow_blank=True),
                col="N"
            )
        )

        for name, d in validation.items():
            d["dv"].error = "Your entry is not in the list"
            d["dv"].errorTitle = 'Invalid Entry'
            ws.add_data_validation(d["dv"])
            for row in range(start_row, last_row + 1):
                d["dv"].add(ws['{0}{1}'.format(d["col"], row)])

        # ***************
        # Filling Columns
        # ***************

        # Adding configured series and their parameters
        for i in range(configured_nb):
            row = i + start_row
            ws['A{}'.format(row)].value = unitcleaners[i].external_name
            ws['B{}'.format(row)].value = unitcleaners[i].name
            # Column C = Save Config, always empty when the Excel is generated
            ws['D{}'.format(row)].value = 'yes'
            ws['E{}'.format(row)].value = unitcleaners[i].freq
            ws['F{}'.format(row)].value = unitcleaners[i].input_unit_type
            ws['G{}'.format(row)].value = unitcleaners[i].input_convention
            ws['H{}'.format(row)].value = unitcleaners[i].clock
            ws['I{}'.format(row)].value = unitcleaners[i].timezone
            ws['J{}'.format(row)].value = unitcleaners[i].unit
            ws['K{}'.format(row)].value = unitcleaners[i].label
            ws['L{}'.format(row)].value = 'yes' if \
                unitcleaners[i].input_expected_regular else 'no'  # boolean could be better managed.
            ws['M{}'.format(row)].value = unitcleaners[i].unit_type
            ws['N{}'.format(row)].value = unitcleaners[i].resample_rule
            ws['O{}'.format(row)].value = unitcleaners[i].interpolate_limit
            ws['P{}'.format(row)].value = unitcleaners[i].wait_offset
            ws['Q{}'.format(row)].value = unitcleaners[i].operation_fct
            ws['R{}'.format(row)].value = unitcleaners[i].filter_fct
            ws['S{}'.format(row)].value = unitcleaners[i].derivative_filter_fct
            ws['T{}'.format(row)].value = unitcleaners[i].custom_delay
            ws['U{}'.format(row)].value = unitcleaners[i].custom_fct
            ws['V{}'.format(row)].value = unitcleaners[i].custom_before_offset
            ws['W{}'.format(row)].value = unitcleaners[i].custom_after_offset

        # Adding non-configured series external names
        for i in range(non_configured_nb):
            row = i + start_row + configured_nb
            ws['A{}'.format(row)].value = not_configured_series[i].external_name

        # *********
        # Styling + Conditional Formatting
        # *********

        green_fill = PatternFill(start_color='66CC99', end_color='66CC99', fill_type='solid')

        center_align = Alignment(
            horizontal='center',
            vertical='center',
            text_rotation=0,
            wrap_text=False,
            shrink_to_fit=False,
            indent=0)

        default_border = Border(
            left=Side(border_style=None, color='FF000000'),
            right=Side(border_style=None, color='FF000000'),
            top=Side(border_style='thin', color='b0b0b0'),
            bottom=Side(border_style='thin', color='b0b0b0')
        )

        right_thin_border = Border(
            right=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='b0b0b0')
        )
        bottom_thin_border = Border(bottom=Side(border_style='thin', color='000000'))
        right_bottom_thin_border = Border(
            bottom=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000')
        )

        for row in range(start_row, last_row + 1):
            for col in range(start_col, last_col + 1):
                # Style
                cell = ws.cell(row=row, column=col)
                cell.border = default_border
                cell.alignment = center_align
                if col == 3:
                    cell.font = Font(bold=True)
                if col != last_col and row == last_row:
                    cell.border = bottom_thin_border
                elif col == last_col and row != last_row:
                    cell.border = right_thin_border
                elif col == last_col and row == last_row:
                    cell.border = right_bottom_thin_border
                # Formatting
                # ws.conditional_formatting.add(
                #     '{0}{1}'.format(cell.column, cell.row),
                #     FormulaRule(
                #         formula=['C{}="x"'.format(row)],
                #         fill=green_fill)
                # )

        # **********
        # write book
        # **********
        wb.save(path)


def excel_to_unitcleaners_data(path, max_input_length=20000):
    unitcleaners_data = []
    wb = load_workbook(path)
    ws = wb["Series"]
    start_row = 8
    for row in range(start_row, max_input_length + start_row):
        # print(ws['C{}'.format(row)].value)
        if ws['C{}'.format(row)].value == 'x':
            if ws['D{}'.format(row)].value == 'yes':
                ws['D{}'.format(row)].value = 'true'
            elif ws['D{}'.format(row)].value == 'no':
                ws['D{}'.format(row)].value = 'false'
            unitcleaners_data.append(dict(
                # An empty cell is None
                external_name=ws['A{}'.format(row)].value,
                name=ws['B{}'.format(row)].value if ws['B{}'.format(row)].value else ws['A{}'.format(row)].value,
                freq=ws['E{}'.format(row)].value,
                input_unit_type=ws['F{}'.format(row)].value,
                input_convention=ws['G{}'.format(row)].value,
                clock=ws['H{}'.format(row)].value,
                timezone=ws['I{}'.format(row)].value,
                unit=ws['J{}'.format(row)].value if ws['J{}'.format(row)].value is not None else '',
                label=ws['K{}'.format(row)].value if ws['K{}'.format(row)].value is not None else '',
                input_expected_regular= ws['L{}'.format(row)].value == 'yes', #Boolean could be better managed.
                unit_type=ws['M{}'.format(row)].value if ws['M{}'.format(row)].value is not None else ws['F{}'.format(row)].value,
                resample_rule=ws['N{}'.format(row)].value if ws['N{}'.format(row)].value is not None else 'mean',
                interpolate_limit=ws['O{}'.format(row)].value if ws['O{}'.format(row)].value is not None else 0,
                wait_offset=ws['P{}'.format(row)].value if ws['P{}'.format(row)].value is not None else '6H',
                operation_fct=ws['Q{}'.format(row)].value,
                filter_fct=ws['R{}'.format(row)].value,
                derivative_filter_fct=ws['S{}'.format(row)].value,
                custom_delay=ws['T{}'.format(row)].value,
                custom_fct=ws['U{}'.format(row)].value,
                custom_before_offset=ws['V{}'.format(row)].value,
                custom_after_offset=ws['W{}'.format(row)].value
            ))

    return unitcleaners_data
