# -*- coding: utf-8 -*-
# This file is a part of DDT (https://github.com/juewuer/ddt2)
# Copyright 2012-2020 Juewuer and DDT contributors
# For the exact contribution history, see the git revision log.
# DDT2 is licensed under the MIT License, included in
# https://github.com/juewuer/ddt2/blob/master/LICENSE.md

import inspect
import json
import os
import re
import codecs
import copy
from functools import wraps

try:
    import yaml
except ImportError:  # pragma: no cover
    _have_yaml = False
else:
    _have_yaml = True

__version__ = '11.3.0.4'

# These attributes will not conflict with any real python attribute
# They are added to the decorated test method and processed later
# by the `ddt` class decorator.

DATA_ATTR = '%values'              # store the data the test must run with
FILE_ATTR = '%file_path'           # store the path to JSON file
FILE_EXTR_ATTR = 'Sheet1'      # for excel, store sheet name
YAML_LOADER_ATTR = '%yaml_loader'  # store custom yaml loader for serialization
UNPACK_ATTR = '%unpack'            # remember that we have to unpack values
AUTOINDEX_ATTR = '%autoindex'            # remember that we have to unpack values
index_len = 5                      # default max length of case index


try:
    trivial_types = (type(None), bool, int, float, basestring)
except NameError:
    trivial_types = (type(None), bool, int, float, str)


def is_trivial(value):
    if isinstance(value, trivial_types):
        return True
    elif isinstance(value, (list, tuple)):
        return all(map(is_trivial, value))
    return False


def unpack(func):
    """
    Method decorator to add unpack feature.

    """
    setattr(func, UNPACK_ATTR, True)
    return func
    
def autoindex(func):
    """
    Method decorator to add unpack feature.

    """
    setattr(func, AUTOINDEX_ATTR, True)
    return func

def data(*values):
    """
    Method decorator to add to your test methods.

    Should be added to methods of instances of ``unittest.TestCase``.

    """
    global index_len
    index_len = len(str(len(values)))
    return idata(values)


def idata(iterable):
    """
    Method decorator to add to your test methods.

    Should be added to methods of instances of ``unittest.TestCase``.

    """
    def wrapper(func):
        setattr(func, DATA_ATTR, iterable)
        return func
    return wrapper


def file_data(value, yaml_loader=None, sheet="Sheet1"):
    """
    Method decorator to add to your test methods.

    Should be added to methods of instances of ``unittest.TestCase``.

    ``value`` should be a path relative to the directory of the file
    containing the decorated ``unittest.TestCase``. The file
    should contain JSON encoded data, that can either be a list or a
    dict.

    In case of a list, each value in the list will correspond to one
    test case, and the value will be concatenated to the test method
    name.

    In case of a dict, keys will be used as suffixes to the name of the
    test case, and values will be fed as test data.

    ``yaml_loader`` can be used to customize yaml deserialization.
    The default is ``None``, which results in using the ``yaml.safe_load``
    method.
    """
    def wrapper(func):
        if value.endswith('xlsx'):
            setattr(func, FILE_EXTR_ATTR, sheet)
            print(sheet, func)
        setattr(func, FILE_ATTR, value)
        if yaml_loader:
            setattr(func, YAML_LOADER_ATTR, yaml_loader)
        return func
    return wrapper


def mk_test_name(name, value, index=0,func=None):
    """
    Generate a new name for a test case.

    It will take the original test name and append an ordinal index and a
    string representation of the value, and convert the result into a valid
    python identifier by replacing extraneous characters with ``_``.

    We avoid doing str(value) if dealing with non-trivial values.
    The problem is possible different names with different runs, e.g.
    different order of dictionary keys (see PYTHONHASHSEED) or dealing
    with mock objects.
    Trivial scalar values are passed as is.

    A "trivial" value is a plain scalar, or a tuple or list consisting
    only of trivial values.
    """

    # Add zeros before index to keep order
    if func is None:
        index = "{0:0{1}}".format(index + 1, index_len)
        if not is_trivial(value):
            return "{0}_{1}".format(name, index)
        try:
            value = str(value)
        except UnicodeEncodeError:
            # fallback for python2
            value = value.encode('ascii', 'backslashreplace')
        test_name = "{0}_{1}_{2}".format(name, index, value)
        test_name = re.sub(r'\W|^(?=\d)', '_', test_name)
        test_name = re.sub(r'_+', '_', test_name)
        test_name = re.sub(r'_$', '', test_name)
        return test_name
    else:
        index = "{0:0{1}}".format(index + 1, index_len)
        if type(value) == type([]) or type(value) == type(()):
            values_default = copy.deepcopy(list(value))
        else:
            values_default = [value]
        values_name = list(func.__code__.co_varnames[1:func.__code__.co_argcount])
        if not is_trivial(value):
            return "{0}_{1}".format(name, index)
        try:
            value = str(value)
        except UnicodeEncodeError:
            # fallback for python2
            value = value.encode('ascii', 'backslashreplace')
        
        #print(values_default,len(values_default),func.__code__.co_argcount,func.__code__.co_varnames)
        if len(values_default)+1 < func.__code__.co_argcount:
            values_default.extend(func.__defaults__[func.__code__.co_argcount-1-len(values_default):])
        values_name.append("i")
        values_default.append(index)
        #print("values_default: ",values_default)
        #print(list(enumerate(func.__code__.co_varnames[1:func.__code__.co_argcount])))
        test_name = name
        founded = False
            
        for i, var in enumerate(values_name):
            #print("to test  __%s__"%(str(var)),"__%s__"%(str(var)) in test_name)
            holder = "__%s__"%(str(var))
            if holder in test_name:
                #print("get %s, set to %s"%( holder, str(values_default[i])) )
                test_name = test_name.replace(holder,str(values_default[i]))
                founded = True
        #print("if founded: ",founded)
        if hasattr(func, AUTOINDEX_ATTR) and "__i__" not in name:
            test_name += "_%s"%str(index)
        elif not founded:
            test_name = "{0}_{1}_{2}".format(test_name, index, value)

            
        #print(test_name, (name, index, value))
        test_name = re.sub(r'\W|^(?=\d)', '_', test_name)
        test_name = re.sub(r'_+', '_', test_name)
        test_name = re.sub(r'_$', '', test_name)
        return test_name    


def feed_data(func, new_name, test_data_docstring, *args, **kwargs):
    """
    This internal method decorator feeds the test data item to the test.

    """
    @wraps(func)
    def wrapper(self):
        return func(self, *args, **kwargs)
    wrapper.__name__ = new_name
    wrapper.__wrapped__ = func
    # set docstring if exists
    if test_data_docstring is not None:
        wrapper.__doc__ = test_data_docstring
    else:
        # Try to call format on the docstring
        if func.__doc__:
            try:
                wrapper.__doc__ = func.__doc__.format(*args, **kwargs)
            except (IndexError, KeyError):
                # Maybe the user has added some of the formating strings
                # unintentionally in the docstring. Do not raise an exception
                # as it could be that user is not aware of the
                # formating feature.
                pass
    return wrapper


def add_test(cls, test_name, test_docstring, func, *args, **kwargs):
    """
    Add a test case to this class.

    The test will be based on an existing function but will give it a new
    name.

    """
    if hasattr(cls, test_name):
        raise ValueError("%s have already have test function %s" % (cls.__name__, test_name))
    setattr(cls, test_name, feed_data(func, test_name, test_docstring,
            *args, **kwargs))


def process_file_data(cls, name, func, file_attr):
    """
    Process the parameter in the `file_data` decorator.
    """
    cls_path = os.path.abspath(inspect.getsourcefile(cls))
    data_file_path = os.path.join(os.path.dirname(cls_path), file_attr)

    def create_error_func(message):  # pylint: disable-msg=W0613
        def func(*args):
            raise ValueError(message % file_attr)
        return func

    # If file does not exist, provide an error function instead
    if not os.path.exists(data_file_path):
        test_name = mk_test_name(name, "error")
        test_docstring = """Error!"""
        add_test(cls, test_name, test_docstring,
                 create_error_func("%s does not exist"), None)
        return

    _is_yaml_file = data_file_path.endswith((".yml", ".yaml"))

    # Don't have YAML but want to use YAML file.
    if _is_yaml_file and not _have_yaml:
        test_name = mk_test_name(name, "error")
        test_docstring = """Error!"""
        add_test(
            cls,
            test_name,
            test_docstring,
            create_error_func("%s is a YAML file, please install PyYAML"),
            None
        )
        return
    if data_file_path.endswith((".xlsx")):
        from openpyxl import load_workbook
        sheetName= getattr(func, FILE_EXTR_ATTR)
        wb = load_workbook(data_file_path)
        print("Currentsheet: ", sheetName)
        ws = wb[sheetName]
        rowNum = ws.max_row
        colNum = ws.max_column
        keys = next(ws.values)
        values = list(ws.values)[1:]
        if rowNum > 1:
            data = []
            for row in list(range(2, rowNum+1)):
                if not any(values[row-2]):
                    continue
                s = {}
                s['rowNum'] = row
                for i, key in enumerate(keys):
                    s[keys[i]] = values[row-2][i]
                data.append(s)
    else:
        with codecs.open(data_file_path, 'r', 'utf-8') as f:
            # Load the data from YAML or JSON
            if _is_yaml_file:
                if hasattr(func, YAML_LOADER_ATTR):
                    yaml_loader = getattr(func, YAML_LOADER_ATTR)
                    data = yaml.load(f, Loader=yaml_loader)
                else:
                    data = yaml.safe_load(f)
            else:
                data = json.load(f)

    _add_tests_from_data(cls, name, func, data)


def _add_tests_from_data(cls, name, func, data):
    """
    Add tests from data loaded from the data file into the class
    """
    for i, elem in enumerate(data):
        if isinstance(data, dict):
            key, value = elem, data[elem]
            test_name = mk_test_name(name, key, i)
        elif isinstance(data, list):
            value = elem
            test_name = mk_test_name(name, value, i)
        if isinstance(value, dict):
            add_test(cls, test_name, test_name, func, **value)
        else:
            add_test(cls, test_name, test_name, func, value)


def _is_primitive(obj):
    """Finds out if the obj is a "primitive". It is somewhat hacky but it works.
    """
    return not hasattr(obj, '__dict__')


def _get_test_data_docstring(func, value):
    """Returns a docstring based on the following resolution strategy:
    1. Passed value is not a "primitive" and has a docstring, then use it.
    2. In all other cases return None, i.e the test name is used.
    """
    if not _is_primitive(value) and value.__doc__:
        return value.__doc__
    else:
        return None


def ddt(cls):
    """
    Class decorator for subclasses of ``unittest.TestCase``.

    Apply this decorator to the test case class, and then
    decorate test methods with ``@data``.

    For each method decorated with ``@data``, this will effectively create as
    many methods as data items are passed as parameters to ``@data``.

    The names of the test methods follow the pattern
    ``original_test_name_{ordinal}_{data}``. ``ordinal`` is the position of the
    data argument, starting with 1.

    For data we use a string representation of the data value converted into a
    valid python identifier.  If ``data.__name__`` exists, we use that instead.

    For each method decorated with ``@file_data('test_data.json')``, the
    decorator will try to load the test_data.json file located relative
    to the python file containing the method that is decorated. It will,
    for each ``test_name`` key create as many methods in the list of values
    from the ``data`` key.

    """
    for name, func in list(cls.__dict__.items()):
        if hasattr(func, DATA_ATTR):
            for i, v in enumerate(getattr(func, DATA_ATTR)):
                test_name = mk_test_name(name, getattr(v, "__name__", v), i, func)
                test_data_docstring = _get_test_data_docstring(func, v)
                if hasattr(func, UNPACK_ATTR):
                    if isinstance(v, tuple) or isinstance(v, list):
                        add_test(
                            cls,
                            test_name,
                            test_data_docstring,
                            func,
                            *v
                        )
                    else:
                        # unpack dictionary
                        add_test(
                            cls,
                            test_name,
                            test_data_docstring,
                            func,
                            **v
                        )
                else:
                    add_test(cls, test_name, test_data_docstring, func, v)
            delattr(cls, name)
        elif hasattr(func, FILE_ATTR):
            file_attr = getattr(func, FILE_ATTR)
            process_file_data(cls, name, func, file_attr)
            delattr(cls, name)
    return cls
